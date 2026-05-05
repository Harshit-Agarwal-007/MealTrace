# app/services/report_service.py
"""
Report generation service — Excel exports for admin dashboard.

Report Types:
  1. Weekly Attendance Report
  2. Financial / Payment Transaction Log
  3. Residents roster
  4. Scan activity (successful scans, date range)
"""

import io
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, headers: list):
    """Apply consistent header styling."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 5)


def generate_weekly_report(start_date: Optional[datetime] = None) -> bytes:
    """
    Weekly attendance report — shows daily scan counts per resident per site.

    Columns: Date | Resident | Site | Breakfast | Lunch | Dinner | Total
    """
    db = get_db()

    if start_date is None:
        start_date = datetime.now(timezone.utc) - timedelta(days=7)

    end_date = start_date + timedelta(days=7)

    # Fetch all scans in the period (uses default single-field index on 'timestamp')
    all_logs_raw = (
        db.collection("scan_logs")
        .where("timestamp", ">=", start_date)
        .where("timestamp", "<=", end_date)
        .get()
    )
    # Filter for SUCCESS client-side
    logs = [d for d in all_logs_raw if d.to_dict().get("status") == "SUCCESS"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Attendance"

    headers = ["Date", "Resident ID", "Resident Name", "Site", "Breakfast", "Lunch", "Dinner", "Total Meals"]
    _style_header(ws, headers)

    # Aggregate by date + resident
    daily_data = {}
    for doc in logs:
        data = doc.to_dict()
        ts = data.get("timestamp")
        if hasattr(ts, "date"):
            day = ts.date().isoformat()
        else:
            day = str(ts)[:10]

        key = (day, data.get("resident_id", ""))
        if key not in daily_data:
            daily_data[key] = {"breakfast": 0, "lunch": 0, "dinner": 0, "site": data.get("site_id", "")}

        meal = (data.get("meal_type") or "").strip().lower()
        if meal in daily_data[key]:
            daily_data[key][meal] += 1

    # Fetch resident names
    resident_names = {}
    for (day, rid), counts in daily_data.items():
        if rid not in resident_names:
            rdoc = db.collection("residents").document(rid).get()
            resident_names[rid] = rdoc.to_dict().get("name", rid) if rdoc.exists else rid

    row = 2
    for (day, rid), counts in sorted(daily_data.items()):
        total = counts["breakfast"] + counts["lunch"] + counts["dinner"]
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=rid)
        ws.cell(row=row, column=3, value=resident_names.get(rid, rid))
        ws.cell(row=row, column=4, value=counts["site"])
        ws.cell(row=row, column=5, value=counts["breakfast"])
        ws.cell(row=row, column=6, value=counts["lunch"])
        ws.cell(row=row, column=7, value=counts["dinner"])
        ws.cell(row=row, column=8, value=total)
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_financial_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> bytes:
    """
    Financial / payment transaction log.

    Optional ``start_date`` / ``end_date`` (inclusive of whole end day when passed from API)
    filter by payment ``timestamp``. When both are omitted, all payments are included (newest first).

    Columns: Date | Resident | Plan | Amount | Status | Razorpay ID
    """
    db = get_db()
    coll = db.collection("payments")

    if start_date is None and end_date is None:
        payments = list(coll.order_by("timestamp", direction="DESCENDING").get())
    else:
        q = coll
        if start_date is not None:
            q = q.where("timestamp", ">=", start_date)
        if end_date is not None:
            q = q.where("timestamp", "<=", end_date)
        raw = list(q.get())
        payments = sorted(
            raw,
            key=lambda d: d.to_dict().get("timestamp") or datetime.min.replace(tzinfo=timezone.utc),
            reverse=True,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    headers = ["Date", "Resident ID", "Plan", "Amount (₹)", "Status", "Razorpay Order ID", "Payment ID"]
    _style_header(ws, headers)

    row = 2
    for doc in payments:
        data = doc.to_dict()
        ts = data.get("timestamp")
        ws.cell(row=row, column=1, value=str(ts)[:19] if ts else "")
        ws.cell(row=row, column=2, value=data.get("resident_id", ""))
        ws.cell(row=row, column=3, value=data.get("plan_id", ""))
        ws.cell(row=row, column=4, value=data.get("amount", 0))
        ws.cell(row=row, column=5, value=data.get("status", ""))
        ws.cell(row=row, column=6, value=data.get("razorpay_order_id", ""))
        ws.cell(row=row, column=7, value=data.get("razorpay_payment_id", ""))
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_residents_report() -> bytes:
    """
    Full residents roster export.

    Columns: ID | Name | Email | Phone | Room | Site | Status | Balance | Plan | Dietary | Created
    """
    db = get_db()

    residents = db.collection("residents").get()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Residents Roster"

    headers = [
        "Resident ID", "Name", "Email", "Phone", "Room Number",
        "Site ID", "Status", "Balance", "Plan ID", "Dietary Preference", "Created At"
    ]
    _style_header(ws, headers)

    row = 2
    for doc in residents:
        data = doc.to_dict()
        ts = data.get("created_at")
        ws.cell(row=row, column=1, value=doc.id)
        ws.cell(row=row, column=2, value=data.get("name", ""))
        ws.cell(row=row, column=3, value=data.get("email", ""))
        ws.cell(row=row, column=4, value=data.get("phone", ""))
        ws.cell(row=row, column=5, value=data.get("room_number", ""))
        ws.cell(row=row, column=6, value=data.get("site_id", ""))
        ws.cell(row=row, column=7, value=data.get("status", ""))
        ws.cell(row=row, column=8, value=data.get("balance", 0))
        ws.cell(row=row, column=9, value=data.get("plan_id", "") or "")
        ws.cell(row=row, column=10, value=data.get("dietary_preference", "VEG"))
        ws.cell(row=row, column=11, value=str(ts)[:19] if ts else "")
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_scans_activity_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    site_id: Optional[str] = None,
    max_rows: int = 50_000,
) -> bytes:
    """
    Successful scan log export for a time range (status SUCCESS only).

    Scans are written to ``scan_logs`` when vendors validate QR codes; this export
    reviews successful meals beyond the capped admin live feeds.

    Columns: Timestamp | Scan ID | Resident | Site | Vendor | Meal | Status | Block reason | Guest pass | Manual | Notes
    """
    db = get_db()

    if end_date is None:
        end_date = datetime.now(timezone.utc)
    if start_date is None:
        start_date = end_date - timedelta(days=7)

    all_logs_raw = (
        db.collection("scan_logs")
        .where("timestamp", ">=", start_date)
        .where("timestamp", "<=", end_date)
        .get()
    )
    logs = [d for d in all_logs_raw if d.to_dict().get("status") == "SUCCESS"]
    if site_id:
        logs = [d for d in logs if d.to_dict().get("site_id") == site_id]
    logs.sort(
        key=lambda d: d.to_dict().get("timestamp", datetime.min.replace(tzinfo=timezone.utc)),
        reverse=True,
    )
    if len(logs) > max_rows:
        logs = logs[:max_rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scan Activity"

    headers = [
        "Timestamp (UTC)",
        "Scan ID",
        "Resident ID",
        "Site ID",
        "Vendor ID",
        "Meal Type",
        "Status",
        "Block Reason",
        "Guest Pass",
        "Manual Scan",
        "Description",
    ]
    _style_header(ws, headers)

    row = 2
    for doc in logs:
        data = doc.to_dict()
        ts = data.get("timestamp")
        ws.cell(row=row, column=1, value=str(ts)[:26] if ts else "")
        ws.cell(row=row, column=2, value=doc.id)
        ws.cell(row=row, column=3, value=data.get("resident_id", ""))
        ws.cell(row=row, column=4, value=data.get("site_id", ""))
        ws.cell(row=row, column=5, value=data.get("vendor_id", ""))
        ws.cell(row=row, column=6, value=data.get("meal_type", ""))
        ws.cell(row=row, column=7, value=data.get("status", ""))
        ws.cell(row=row, column=8, value=str(data.get("block_reason") or ""))
        ws.cell(row=row, column=9, value="Y" if data.get("is_guest_pass") else "")
        ws.cell(row=row, column=10, value="Y" if data.get("is_manual") else "")
        ws.cell(row=row, column=11, value=str(data.get("description") or ""))
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
